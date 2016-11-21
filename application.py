# coding: utf-8

import os
import pymongo
import logging
import subprocess
import sys
import re
import requests
import shutil
import json
import io
import urllib
import codecs

from subprocess import Popen, check_output
from base64 import b64encode
from flask import Flask, jsonify, request, render_template, redirect, url_for, send_from_directory, send_file
from flask.json import JSONEncoder
from werkzeug import secure_filename
from pymongo import MongoClient
from celery import Celery
from bson.objectid import ObjectId
from raven.contrib.flask import Sentry
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import xlsxwriter
from io import BytesIO
import ipautils

from datetime import datetime
import time

import settings
import global_data

from decorators import crossdomain
from token_gen import Get_Header

app = Flask(__name__, static_url_path='')
app.config['SENTRY_DSN'] = 'SENTRY_KEY'
app.config['CAROUSEL_FOLDER'] = settings.CAROUSEL_FOLDER
app.config['BUNDLES_FOLDER'] = settings.BUNDLES_FOLDER
application_db_con = MongoClient(settings.MONGODB_HOST, settings.MONGODB_PORT)

if not os.environ.get("DEVELOPMENT"):
    sentry = Sentry(app)
    idgen_con = MongoClient(settings.IDGEN_MONGO, settings.MONGODB_PORT)
else:
    idgen_con = MongoClient(settings.MONGODB_HOST, settings.MONGODB_PORT)

appview = application_db_con['appview']
requestsdb = application_db_con['requests']
db = application_db_con['application_db']

apps = db['apps']
groups = db['groups']
repo = db['repo']
users = db['users']

banners = appview['banners']
bundles = appview['bundles']
sections = appview['sections']
views = appview['views']

download = requestsdb['download']

idgen = idgen_con['idgen']
idreq = idgen['idreq']

apps.ensure_index([('nam', pymongo.TEXT)])
banners.ensure_index([('name', pymongo.TEXT)])
users.ensure_index([('udid', pymongo.TEXT), ('name', pymongo.TEXT), ('signature', pymongo.TEXT)])

try:
    capp = Celery(__name__, broker=settings.BROKER_URL)
except:
    logging.error('Could not initiate celery!')
    capp = None

class CustomJSONEncoder(JSONEncoder):
    def default(self, o):
        if isinstance(o, ObjectId):
            return str(o)

        return JSONEncoder.default(self, o)

app.json_encoder = CustomJSONEncoder

USERS_TOKEN = Get_Header('users')
F5_TOKEN = Get_Header('f5')
####################################################
# UI                                               #
####################################################

@app.route('/')
def index():
    return render_template('index.html')

####################################################
# USERS                                            #
####################################################
@app.route('/users/list')
@crossdomain(origin='*')
def users_list():
    email_reg = re.compile(r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)")
    mobile_reg = re.compile(r'\d{11}')
    page_size = int(request.args.get('count', 12))
    page = int(request.args.get('page', 0))
    query = request.args.get('q', None)

    criteria = {}

    if query:
        data = {}
        if email_reg.match(query):
            header = USERS_TOKEN.get_api_header()
            data['email'] = email_reg.search(query).group(0)
            udid_list = requests.post(settings.SERVICES + "/panel/get_udid", headers=header, data=data).json()
            items = list(users.find({"udid": {"$in": udid_list}}, skip=page*page_size, limit=page_size))
            return jsonify(users=items, total_count=len(items))
        elif mobile_reg.match(query):
            header = USERS_TOKEN.get_api_header()
            data['mobile'] = mobile_reg.search(query).group(0)
            udid_list = requests.post(settings.SERVICES + "/panel/get_udid", headers=header, data=data).json()
            items = list(users.find({"udid": {"$in": udid_list}}, skip=page*page_size, limit=page_size))
            return jsonify(users=items, total_count=len(items))
        else:
            criteria['$text'] = {
                '$search': query
            }

    items = list(users.find(criteria, skip=page*page_size, limit=page_size))
    total_count = users.find(criteria).count()

    return jsonify(users=items, total_count=total_count)

@app.route('/user/get_data', methods=["POST"])
@crossdomain(origin='*')
def user_data():
    """Get a udid from a POST request and return all it's related data using `users` app API.
    The result is a json output separated in 'user', 'device' and 'f5' categeories.
    """
    header = USERS_TOKEN.get_api_header()
    udid = request.form.get('udid', None)
    if udid == None:
        return jsonify(done=False)
    try:
        data = requests.post(settings.SERVICES+'/panel/get_data', headers=header, data={'udid':udid}).json()
    except:
        return jsonify(done=False)
    device_data = data[0]
    user_data = data[1]
    f5_data = data[2]
    return jsonify(device=device_data, user=user_data, f5=f5_data, done=True)


@app.route('/users/block_switch', methods=["POST"])
@crossdomain(origin='*')
def block_device_switch():
    """Get a udid from a POST request and block its related user using `users` app API
    """
    header = USERS_TOKEN.get_api_header()
    udid = request.form.get('udid', None)
    if udid == None:
        return jsonify(done=False)
    try:
        res = requests.post(settings.SERVICES+'/panel/device/block_switch', headers=header, data={'udid': udid})
    except:
        return jsonify(done=False)
    return jsonify(res.json())

@app.route('/users/resend_install_email', methods=["POST"])
@crossdomain(origin='*')
def resend_install_mail():
    udid = request.form.get('udid', None)

    if capp:
        capp.send_task('appsign.udid_tasks.register_udid', args=[udid, ''], queue='udid')
        return jsonify(done=True)
    else:
        return jsonify(done=False)

@app.route('/users/resend_activation_email', methods=["POST"])
@crossdomain(origin='*')
def resend_activation_email():
    email = request.form.get('email', None)
    res = os.system("/path_to_support/first_email.sh {}".format(email))
    return jsonify(done=True if not res else False)

@app.route('/users/block', methods=["POST"])
@crossdomain('*')
def users_block():
    user = request.form.get('udid', None)
    print(user)
    return jsonify(done=True)


####################################################
# GROUPS                                           #
####################################################

@app.route('/groups/list')
@crossdomain(origin='*')
def groups_list():
    page_size = int(request.args.get('count', 12))
    page = int(request.args.get('page', 0))

    items = list(groups.find({}, skip=page*page_size, limit=page_size).sort([('order', -1)]))
    total_count = groups.find({}).count()

    return jsonify(groups=items, total_count=total_count)

@app.route('/groups/add', methods=["POST"])
@crossdomain(origin='*')
def groups_add():
    devid = request.form['devid']
    username = request.form['username']
    password = request.form['password']
    device_count = int(request.form.get('device_count', 0))
    max_devices = int(request.form.get('max_devices', 90))
    order = int(request.form.get('order', 1000))
    enterprise = True if request.form.get('enterprise', '') == 'true' else False
    profile_name = request.form['profile_name']

    # TODO: upload profile file

    gid = groups.insert({
        'devid': devid,
        'username': username,
        'password': password,
        'device_count': device_count,
        'max_devices': max_devices,
        'enterprise': enterprise,
        'order': order,
        'profile_name': profile_name,
        'full': False
    })

    return jsonify(id=gid, done=True)

@app.route('/groups/edit', methods=["POST"])
@crossdomain(origin='*')
def groups_edit():
    group_id = request.form['id']
    devid = request.form['devid']
    username = request.form['username']
    password = request.form['password']
    device_count = int(request.form.get('device_count', 0))
    max_devices = int(request.form.get('max_devices', 90))
    order = int(request.form.get('order', 1000))
    enterprise = True if request.form.get('enterprise', '') == 'true' else False
    profile_name = request.form['profile_name']

    groups.update({'_id': ObjectId(group_id)}, {'$set': {
        'devid': devid,
        'username': username,
        'password': password,
        'device_count': device_count,
        'max_devices': max_devices,
        'enterprise': enterprise,
        'profile_name': profile_name,
        'order': order
    }}, upsert=False, multi=False)

    return jsonify(done=True)

@app.route('/groups/assign_user', methods=["POST"])
@crossdomain(origin='*')
def groups_assign_user():
    group_id = request.form['group_id']
    udid = request.form['udid']
    return jsonify(done=True)


####################################################
# APPS                                             #
####################################################

@app.route('/apps/list')
@crossdomain(origin='*')
def apps_list():
    page_size = int(request.args.get('count', 12))
    page = int(request.args.get('page', 0))
    disabled = True if request.args.get('disabled', '') == 'true' else False
    query = request.args.get('q')

    criteria = {}

    if disabled:
        criteria = {
            'status': {'$gte': 100},
        }
    else:
        criteria = {
            'status': {'$lt': 100},
            'cop': {
                '$elemMatch':
                {'status': {'$lt': 100}}
            }
        }

    if query:
        criteria['$text'] = {
            '$search': query
        }

        items = list(apps.find(criteria, {'score': {'$meta': "textScore"}}, skip=page * page_size, limit=page_size).sort([('score', {'$meta': 'textScore'})]))
    else:
        items = list(apps.find(criteria, skip=page * page_size, limit=page_size))

    total_count = apps.find(criteria).count()

    return jsonify(apps=items, total_count=total_count)

@app.route('/apps/find')
@crossdomain(origin='*')
def find_apps():
    term = request.args.get('q')
    try:
        data = requests.get(settings.API_FIND, params={'q': term})
        res = data.json()
    except:
        res = {}
    return jsonify(res)

@app.route('/appview/banners')
@crossdomain(origin='*')
def banners_list():
    banner_type = request.args.get('type', 'banners')
    page_size = int(request.args.get('count', 12))
    page = int(request.args.get('page', 0))
    query = request.args.get('q', None)

    criteria = {}
    if query:
        criteria['$text'] = {
            '$search': query
        }
    items = list(appview[banner_type].find(criteria,{'score': {'$meta': "textScore"}}, skip=page * page_size, limit=page_size))
    count = len(items)
    return jsonify(banners=items, count=count)


def uri_verifier(uri_list):
    """
    Get a uri list and return 3 set of magnet, torrent and regular http uris
    """
    web_uri = set()
    magnet_uri = set()
    torrent_uri = set()
    invalid_uri = set()

    # Regex to match a url:
    # http://stackoverflow.com/a/7160778/1573477
    web_regex = re.compile(
        r'^(?:http|ftp)s?://' # http:// or https://
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' #domain...
        r'localhost|' #localhost...
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})' # ...or ip
        r'(?::\d+)?' # optional port
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)

    for uri in uri_list:
        if uri.startswith('magnet:'):
            magnet_uri.add(uri)
        elif uri.endswith('torrent'):
            torrent_uri.add(uri)
        elif web_regex.match(uri):
            web_uri.add(uri)
        else:
            invalid_uri.add(uri)

    return {'webs': web_uri, 'magnets': magnet_uri, 'invalids': invalid_uri}

@app.route('/apps/add', methods=["POST"])
@crossdomain(origin='*')
def apps_add():
    uri_list = request.form.get('url', '').replace('\n', '').replace(' ', '').split(',')

    checked_urls = uri_verifier(uri_list)

    valid_urls = []
    invalid_urls = list(checked_urls['invalids'])
    for url in checked_urls['webs']:
        valid_urls.append(url)
        link = [[url, True], ]
        query = {'id': ObjectId(), 'applinkid': None, 'canDownload': 1, 'data': {},
                 'links': link, 'report': True, 'starred': True}
        requestsdb.download.insert(query)
    for magnet in checked_urls['magnets']:
        valid_urls.append(magnet)
        process = Popen(['bash', 'scripts/magnet2torrent.sh', magnet, settings.TORRENT_PATH])

    return jsonify(done=True, invalid_urls=invalid_urls, valid_urls=valid_urls)

@app.route('/apps/add/by_link')
@crossdomain(origin='*')
def apps_add_by_link():
    link = request.form['link']
    return jsonify(done=True)

@app.route('/apps/add/by_ipa', methods=["POST"])
@crossdomain(origin='*')
def apps_add_by_ipa():
    return jsonify(done=True)

@app.route('/apps/remove', methods=["POST"])
@crossdomain(origin='*')
def apps_remove():
    app_id = request.form.get('id', None)
    if not app_id:
        return jsonify(done=False)

    app = apps.find_one({"_id": ObjectId(app_id)})
    print(app['id'])
    # Remove all cop files and their repo documents.
    for cop in app['cop']:
        delete_repo_and_path(cop['lid'])
    apps.remove(ObjectId(app_id))
    return jsonify(done=True)


def delete_repo_and_path(lid):
    """Get an lid and remove it's related file and repo document field.
    """
    query = {'localid': lid}
    repo_doc = repo.find_one(query)
    cache_path = os.path.join(settings.LAYER3_WORKING_DIR, lid)
    path = repo_doc['path']
    # Remove cop file if exists
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=True)
    if os.path.exists(cache_path):
        shutil.rmtree(cache_path, ignore_errors=True)
    repo.remove(query)

@app.route('/apps/cop_remove', methods=["POST"])
@crossdomain(origin="*")
def cop_remove():
    """Get an lid and remove related cop entry and its file from repos path.
    """
    lid = request.form.get('lid', None)
    if not lid:
        return jsonify(done=False)

    query = {'cop': {'$elemMatch': {'lid': lid}}}
    apps.update(query, {'$pull': {'cop': {'lid': lid}}})
    delete_repo_and_path(lid)
    return jsonify(done=True)

@app.route('/apps/change_status', methods=["POST"])
@crossdomain(origin="*")
def change_status():
    app_id = request.form.get('app_id', None)
    value = request.form.get('value', None)
    if not app_id and value:
        return jsonify(done=False)

    apps.update({"_id": ObjectId(app_id)}, {"$set": {"status": int(value)}})
    return jsonify(done=True)

@app.route('/apps/change_cop_status', methods=["POST"])
@crossdomain(origin="*")
def change_cop_status():
    lid = request.form.get('lid', None)
    value = request.form.get('value', None)
    if not lid and value:
        return jsonify(done=False)

    apps.update({'cop': {'$elemMatch': {'lid': lid}}}, {'$set': {'cop.$.status': int(value)}})
    return jsonify(done=True)

@app.route('/apps/set_app_note', methods=["POST"])
@crossdomain(origin="*")
def set_app_note():
    note = request.form.get('note', None)
    if not note:
        return jsonify(done=False)

    # Check if it's an application set note request it's for a cop.
    if 'id' in request.form.keys():
        app_id = request.form.get('id', None)
        apps.update({"_id": ObjectId(app_id)}, {"$set": {"note": note}})
        return jsonify(done=True)
    elif 'lid' in request.form.keys():
        lid = request.form.get('lid', None)
        apps.update({'cop': {'$elemMatch': {'lid': lid}}}, {'$set': {'cop.$.note': note}})
        return jsonify(done=True)

    return jsonify(done=False)


@app.route('/apps/set_app_desc', methods=["POST"])
@crossdomain(origin="*")
def set_app_desc():
    """
    Get a localized description `desc` for an app and set it using app_id
    """
    desc = request.form.get('desc', None)
    app_id = request.form.get('id', None)
    tags = request.form.get('tags', None)
    if not (desc or tags) or not app_id:
        return jsonify(done=False)

    if tags:
        # Remove non alphanumeric characters from tags
        pattern = re.compile('\W')
        tags = pattern.sub(' ', tags).lower().split()
    else:
        tags = []

    data = {"locdes": desc, "tags": tags}
    apps.update({"_id": ObjectId(app_id)}, {"$set": data})
    return jsonify(done=True)


@app.route('/apps/downloader_log')
@crossdomain(origin="*")
def get_downloader_log():
    try:
        cmd_res = check_output(['/last_axel_change.sh']).decode('ascii')
        log = check_output(['/download_status.sh']).decode('ascii')
        time = re.sub(' +', ' ', cmd_res[:-1])
    except subprocess.CalledProcessError as e:
        return jsonify(err=str(e), done=False)

    return jsonify(log=log.split(' '), time=time.split(' '), done=True)


@app.route('/apps/reset_downloader')
@crossdomain(origin="*")
def reset_downloader():
    try:
        process = check_output(['/reset_downloader.sh'])
    except subprocess.CalledProcessError as e:
        return jsonify(err=str(e), done=False)

    return jsonify(done=True)

def allowed_file(filename):
    allowed_extensions = set(['txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'])
    return '.' in filename and filename.rsplit('.', 1)[1] in allowed_extensions

@app.route('/apps/upload_banner', methods=['POST', 'GET'])
def upload_banner():
    if request.method == 'POST':
        alt = request.form.get('alt', None)
        appid = request.form.get('app', None)
        iphone1 = request.files.get('iphone1', None)
        iphone2 = request.files.get('iphone2', None)
        iphone3 = request.files.get('iphone3', None)
        ipad = request.files.get('ipad@2x', None)
        if alt and iphone1 and iphone2 and iphone3 and ipad and \
           allowed_file(iphone1.filename) and allowed_file(iphone2.filename) and \
            allowed_file(iphone3.filename) and allowed_file(ipad.filename):
            iphone1_img = secure_filename(iphone1.filename)
            iphone2_img = secure_filename(iphone2.filename)
            iphone3_img = secure_filename(iphone3.filename)
            ipad_img = secure_filename(ipad.filename)
            iphone1.save(os.path.join(app.config['CAROUSEL_FOLDER'], iphone1_img))
            iphone2.save(os.path.join(app.config['CAROUSEL_FOLDER'], iphone2_img))
            iphone3.save(os.path.join(app.config['CAROUSEL_FOLDER'], iphone3_img))
            ipad.save(os.path.join(app.config['CAROUSEL_FOLDER'], ipad_img))
 
            img_dict = {"iphone@2x": settings.API_MEDIA + url_for('uploaded_banner', filename=iphone2_img),
                        "iphone-375w@2x": settings.API_MEDIA + url_for('uploaded_banner', filename=iphone1_img),
                        "iphone-320w@2x": settings.API_MEDIA + url_for('uploaded_banner', filename=iphone2_img),
                        "iphone-414w@3x": settings.API_MEDIA + url_for('uploaded_banner', filename=iphone3_img),
                        "ipad@2x": settings.API_MEDIA + url_for('uploaded_banner', filename=ipad_img)}

            query = {"alt": alt.strip()}
            banners.update(query,
                           {"alt": alt.strip(),
                            "app": appid,
                            "image": img_dict},
                           upsert=True)

            itemid = banners.find_one(query)['_id']
            return jsonify(iphone1=url_for('uploaded_banner', filename=iphone1_img),
                           iphone2=url_for('uploaded_banner', filename=iphone2_img),
                           iphone3=url_for('uploaded_banner', filename=iphone3_img),
                           ipad=url_for('uploaded_banner', filename=ipad_img),
                           alt=alt, app=appid, _id=itemid, done=True)

    return jsonify(done=False)

@app.route('/apps/upload_bundle', methods=['POST', 'GET'])
def upload_bundle():
    if request.method == 'POST':
        slug = request.form.get('slug', None)
        title = request.form.get('title', None)
        img1 = request.files.get('img1', None)
        img2 = request.files.get('img2', None)
        img3 = request.files.get('img3', None)
        img4 = request.files.get('img4', None)
        apps = request.form.get('apps', None)
        if title and slug and apps and img1 and img2 and img3 and img4 and \
           allowed_file(img1.filename) and allowed_file(img2.filename) and \
            allowed_file(img3.filename) and allowed_file(img4.filename):

            apps = re.sub(r"\s+", "", apps).split(',')

            slug = slug.replace(' ', '')
            link = "PROJECT://apps/collection/{}?title={}".format(slug,
                                                                    urllib.parse.quote(title))

            img1_obj = secure_filename(img1.filename)
            img2_obj = secure_filename(img2.filename)
            img3_obj = secure_filename(img3.filename)
            img4_obj = secure_filename(img4.filename)
            img1.save(os.path.join(app.config['BUNDLES_FOLDER'], img1_obj))
            img2.save(os.path.join(app.config['BUNDLES_FOLDER'], img2_obj))
            img3.save(os.path.join(app.config['BUNDLES_FOLDER'], img3_obj))
            img4.save(os.path.join(app.config['BUNDLES_FOLDER'], img4_obj))

            img_dict = {"@1x": settings.API_MEDIA + url_for('uploaded_bundle', filename=img1_obj),
                        "@2x": settings.API_MEDIA + url_for('uploaded_bundle', filename=img2_obj),
                        "@3x": settings.API_MEDIA + url_for('uploaded_bundle', filename=img3_obj),
                        "site": settings.API_MEDIA + url_for('uploaded_bundle', filename=img4_obj)}

            query = {"title": title.strip()}
            bundles.update(query,
                           {"title": title.strip(),
                            "link": link,
                            "slug": slug,
                            "apps": apps,
                            "image": img_dict},
                           upsert=True)

            itemid = bundles.find_one(query)['_id']
            return jsonify(img1=url_for('uploaded_bundle', filename=img1_obj),
                           title=title, slug=slug, _id=itemid, done=True)

    return jsonify(done=False)


@app.route('/bundles/<filename>')
def uploaded_bundle(filename):
    return send_from_directory(app.config['BUNDLES_FOLDER'], filename)

@app.route('/banners/<filename>')
def uploaded_banner(filename):
    return send_from_directory(app.config['CAROUSEL_FOLDER'], filename)


@app.route('/appview/save_view', methods=['post'])
@crossdomain(origin='*')
def save_view():
    json_object = json.loads(request.form.get('jsonObject', {}))
    if len(json_object) > 0:
        views.insert(json_object)
        return jsonify(done=True)

    return jsonify(done=False)


@app.route('/appview/save_section', methods=['POST'])
@crossdomain(origin='*')
def save_section():
    json_object = json.loads(request.form.get('jsonObject', {}))
    if len(json_object) > 0:
        sections.update({'type': json_object['type'],
                         'name': json_object['name']},
                        json_object,
                        upsert=True)

        return jsonify(done=True)

    return jsonify(done=False)

def items_data(items, sec_type, full_data=True):
    new_items = []
    if sec_type == 'banner' or sec_type == 'bundle':
        projection = {'_id': False, 'apps': False}
        for banner in items:
            section_db = banners if sec_type == 'banner' else bundles
            data = section_db.find_one({'_id': ObjectId(banner)}, projection)
            new_items.append(data)
    else:
        PROJECT_lookup = apps.find({'id': {'$in': items}})
        not_PROJECT = [x for x in items if x not in PROJECT_lookup.distinct('id')]
        new_items = list(PROJECT_lookup)
        if not_PROJECT:
            itunes_lookup = requests.get(settings.API_LOOKUP, params={'appid': ','.join(not_PROJECT)})
            if itunes_lookup.status_code == 200:
                new_items += itunes_lookup.json()['list']

    if not full_data and not sec_type == 'bundle' and not sec_type == 'banner':
        new_items = [x['id'] for x in new_items]

    return new_items


def inject_data(section, full_data):
    sec_type = section['type']
    if sec_type == 'applist_horizontal':
        if section.get('items', None):
            object_details = items_data(section['items'], sec_type, full_data)
            section['items'] = object_details
    elif sec_type == 'banner' or sec_type == 'bundle':
        object_details = items_data(section['items'], sec_type)
        section['items'] = object_details

    return section


def get_sections(sec_id=None, sec_type=None):
    if sec_id:
        result = sections.find({'_id': ObjectId(sec_id)})
    elif sec_type:
        result = sections.find({'type': sec_type})
    else:
        result = sections.find()

    return list(result)


@app.route('/appview/sections', methods=['POST'])
@crossdomain(origin='*')
def return_sections():
    sec_type = request.form.get('type', None)
    if sec_type:
        sec_list = get_sections(sec_type=sec_type)
        for i in range(len(sec_list)):
            sec_list[i] = inject_data(sec_list[i], full_data=True)

        return jsonify(done=True, sections=sec_list)

    else:
        return jsonify(done=False)


@app.route('/appview/current_view', methods=['GET'])
@crossdomain(origin='*')
def return_current_view():
    json = request.args.get('json', None)
    last_view = views.find().sort('_id', -1).limit(1)[0]
    for i in range(len(last_view['value']['sections'])):
        section = get_sections(sec_id=last_view['value']['sections'][i])[0]
        object_details = inject_data(section, full_data=False if json else True)
        last_view['value']['sections'][i] = object_details

    return jsonify(done=True, view=last_view)


@app.route('/statistics', methods=['GET'])
@crossdomain(origin='*')
def get_statistics():
    in_queue = idreq.find({'created': {"$exists": False}, 'error': {"$exists": False}, 'retry': {'$lt': 3}}).count()
    errors = idreq.find({'error': {'$exists': True}, 'created': {"$exists": False}, 'retry': {'$gte': 3}}).count()
    compelete = idreq.find({'created': 1}).count()
    all_reqs = idreq.find().count()

    return jsonify(in_queue=in_queue, errors=errors, compelete=compelete, all_reqs=all_reqs)


@app.route('/promocheck', methods=['GET'])
@crossdomain(origin='*')
def get_promo_status():
    promo = request.args.get("promo", None)
    response = {'error': True}
    if promo:
        url = "http://FPAN_PATH/check_promo/"
        data = {'promo': promo}
        response = requests.post(url, data=data, headers=F5_TOKEN.get_api_header()).json()
    return jsonify(response)


def xl2dict(xlfile):
    # Dictionaries to translate app/cop status to db friendly values
    app_status = {'عادی': 0,
                  'فعال‌، نشانه گذاری شده است': 1,
                  'غیرفعال‌، نشانه گذاری شده است': 100,
                  'موقتا از دسترس خارج شده است': 101,
                  'برای همیشه از دسترس خارج شده است': 102}
    cop_status = {'تست نشده‌، فعال است': 0,
                  'تست شده‌، مشکلی ندارد': 1,
                  'تست نشده‌، غیرفعال است': 100,
                  'تست شده‌، مشکل دارد': 101}
    wb = load_workbook(xlfile)
    sheet = wb['Test']
    rows = sheet.rows
    # Get folder title from A0
    folder = rows[0][0].value
    apps = {folder: []}
    # Iterate over rows and get app data line by line
    for row in rows[2:]:
        app = {}
        app['id'] = row[0].value
        app['nam'] = row[1].value
        app['spname'] = row[2].value
        app['ver'] = row[3].value
        # Translate app/cop status:
        app['cop-state'] = cop_status.get(row[4].value, None)
        app['app-state'] = app_status.get(row[5].value, None)
        app['note'] = row[6].value or ""
        app['lid'] = row[8].value
        # Append app to apps list
        apps[folder].append(app)
    return apps

def apply_app_test_result(app_data):
    app_id = app_data['id']
    app_status = app_data['app-state']
    cop_status = app_data['cop-state']
    cop_ver = app_data['ver']
    cop_note = app_data['note']
    lid = app_data['lid']
    if app_id and app_status:
        apps.update({'id': app_id},
                    {'$set':{'status': int(app_status)}})
    if app_id and lid and cop_status:
        apps.update({'cop': {'$elemMatch': {'lid': lid}}},
                    {'$set': {'cop.$.status': int(cop_status),
                              'cop.$.note': cop_note}})
    return True
    


@app.route('/apps/upload_test', methods=['POST', 'GET'])
def upload_xl():
    fine_row = []
    bad_row = []
    if request.method == 'POST':
        xlfile = request.files.get('xl-test', None)
        apps_data = xl2dict(BytesIO(xlfile.read()))
        folder = list(apps_data)[0]
        for app in apps_data[folder]:
            submit = apply_app_test_result(app)
            if submit == True:
                fine_row.append(app['id'])
            else:
                bad_row.append(app['id'])

        return jsonify(done=True, fine_row=fine_row, bad_row=bad_row)
    else:
        return jsonify(done=False)

@app.route('/apps/test_folders', methods=['GET'])
def get_test_sign_folders():
    signed_dic = {}
    
    for signed in db['test_sign_queue'].find({'signed': 1}):
        signed_dic[str(signed['lid'])]=signed['dir']
    
    dirs = {k:db['test_sign_queue'].count({'dir':k}) for k in set(list(signed_dic.values()))}
    test_count = {k:0 for k in set(list(signed_dic.values()))}
    
    for app in db['apps'].find({}):
        for cop in app['cop']:
            if cop['lid'] in signed_dic and cop['status'] > 0:
                test_count[signed_dic[cop['lid']]]+=1
    
    
    signed_folders = []
    for d in dirs.keys():
        if test_count[d] == 0:
            signed_folders.append(d+' (Not Tested)')
        elif test_count[d] < dirs[d]:
            signed_folders.append(d+' (In Progress)')
    
    
    return jsonify(done=True, dirs=sorted(signed_folders))

@app.route('/apps/get_test_sheet', methods=['GET'])
def generate_xl():
    folder_name = request.args.get('dir')
    folder_name = folder_name.replace(" (Not Tested)","").replace(" (In Progress)","")
    app_list = db['test_sign_queue'].find({'dir': folder_name})
    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out)
    ws = wb.add_worksheet("Test")

    app_status = {'عادی': 0,
                  'فعال‌، نشانه گذاری شده است': 1,
                  'غیرفعال‌، نشانه گذاری شده است': 100,
                  'موقتا از دسترس خارج شده است': 101,
                  'برای همیشه از دسترس خارج شده است': 102}
    cop_status = {'تست نشده‌، فعال است': 0,
                  'تست شده‌، مشکلی ندارد': 1,
                  'تست نشده‌، غیرفعال است': 100,
                  'تست شده‌، مشکل دارد': 101}
    rev_app_status = {v: k for k, v in app_status.items()}
    rev_cop_status = {v: k for k, v in cop_status.items()}
    header_format = wb.add_format({
        'border': 1,
        'bg_color': '#C6EFCE',
        'bold': True,
        'text_wrap': True,
        'valign': 'vcenter',
        'indent': 1,
    })
    columns = {'app_id': 0, 'name': 1, 'sbname': 2, 'version': 3, 'ipa_state':4,
               'app_state': 5, 'test_desc': 6, 'ins_time': 7, 'lid': 8}

    ws.write_string(0,0,folder_name)
    ws.set_column('A:A', 20)
    ws.set_column('B:B', 68)
    ws.set_column('C:C', 30)
    ws.set_column('E:E', 32)
    ws.set_column('F:F', 32)
    ws.set_column('H:H', 30)
    ws.set_column('I:I', 30)
    
    ws.write_string(1, columns['app_id'],'iTunes ID',header_format)
    ws.write_string(1, columns['name'], 'Name',header_format)
    ws.write_string(1, columns['sbname'],'Springboard Name',header_format)
    ws.write_string(1, columns['version'],'IPA version',header_format)
    ws.write_string(1, columns['ipa_state'],'IPA State',header_format)
    ws.write_string(1, columns['app_state'],'App State',header_format)
    ws.write_string(1, columns['test_desc'],'Test Description',header_format)
    ws.write_string(1, columns['ins_time'],'Import Time',header_format)
    ws.write_string(1, columns['lid'],'Local ID',header_format)

    for row in range(2, app_list.count() + 2):
        app = app_list[row - 2]
        lid = app['lid']
        app_data = db.apps.find_one({'cop': {'$elemMatch': {'lid': str(lid)}}})
        cop_data = db.apps.find_one({'cop.lid': str(lid)}, {'_id': 0, 'cop.$': 1})['cop'][0]
        ipa_state = cop_data['status']
        app_state = app_data['status']
        
        app_state = rev_app_status.get(app_state,"")
        ipa_state = rev_cop_status.get(ipa_state,"")
            
        directory = '/'.join([str(lid)[i:i+3] for i in range(0, len(str(lid)), 3)])
        package_name, ipa_info = ipautils.get_ipa_info("/app/afbackend/repo/{}/data/program.ipa".format(directory))

        ipa_display_name = ipa_info.get('CFBundleDisplayName', ipa_info.get('CFBundleName', 'N/A'))
        
        
        ws.write_string(row, columns['app_id'],app_data['id'])
        ws.write_string(row, columns['name'], app_data['nam'])
        ws.write_string(row, columns['sbname'],ipa_display_name)
        ws.write_string(row, columns['version'],cop_data['ver'])
        ws.write_string(row, columns['ipa_state'],ipa_state)
        ws.write_string(row, columns['app_state'],app_state)
        ws.write_string(row, columns['test_desc'],cop_data['note'])
        ws.write_string(row, columns['ins_time'],str(ObjectId(lid).generation_time))
        ws.write_string(row, columns['lid'],str(lid))

    ws.data_validation(2,4,app_list.count()+1,4, {'validate': 'list',
                                  'source': list(cop_status.keys())})
    
    ws.data_validation(2,5,app_list.count()+1,5, {'validate': 'list',
                                  'source': list(app_status.keys())})
    
    
    wb.close()
    out.seek(0)

    return send_file(out,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     attachment_filename='{}.xlsx'.format(folder_name), as_attachment=True)


def upload_file(target, path, name=None):
    filename = secure_filename(target.filename)
    if name:
        filename = ".".join([name, filename.split('.')[-1]])

    target.save(os.path.join(path, filename))

def create_dir(path):
    if not os.path.exists(path):
        os.makedirs(path)

@app.route('/apps/import', methods=['POST'])
def import_apps():
    data = {}
    for item in ['nam', 'cat', 'ven', 'des', 'locdes', 'note']:
        data[item] = request.form.get(item)

    # Set genere id
    data["cid"] = global_data.cat_id[data['cat']]

    # Format price
    prc = re.sub('[^0-9.]', '', request.form.get('prc'))
    try:
        if float(data) == 0:
            raise
        prc += "$"
    except:
        prc = "Free"
    data['prc'] = prc

    # Set supported devices
    sdev = request.form.getlist("sdev")
    if 'All' in sdev:
        sdev = global_data.device_list
    data['sdev'] = sdev

    # Process files
    artwork = request.files.get('artwork')
    ipa = request.files.get("ipa")
    scr = request.files.getlist("iphone")
    bscr = request.files.getlist("ipad")

    # if not (nam and cat and ven and prc and artwork and ipa and scr):
        # return jsonify({'done': False})

    data["rel"] = str(time.mktime(datetime.now().timetuple())).split('.')[0]

    # Create app directory in working path
    appdir = os.path.join(settings.IMPORT_DIR, data['rel'])
    scrdir = os.path.join(appdir, 'iphone')
    bscrdir = os.path.join(appdir, 'ipad')
    create_dir(appdir)

    # Upload files
    upload_file(ipa, appdir, 'program')
    upload_file(artwork, appdir, 'artwork')
    for item in scr:
        if not item.filename:
            continue
        create_dir(scrdir)
        upload_file(item, scrdir)
    for item in bscr:
        if not item.filename:
            continue
        create_dir(bscrdir)
        upload_file(item, bscrdir)

    # Write json data
    with codecs.open(os.path.join(appdir, "data.json"), "w", 'utf8') as datafile:
        datafile.write(json.dumps(data, datafile, ensure_ascii=False))


    return jsonify({'done': True})

if __name__ == '__main__':
    app.debug = True
    app.run()
